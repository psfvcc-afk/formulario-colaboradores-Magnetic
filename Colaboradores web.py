import streamlit as st
import pandas as pd
import dropbox
from datetime import datetime
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

st.set_page_config(
    page_title="Registo de Colaboradores",
    page_icon="📋",
    layout="centered"
)

# Configuração OAuth 2 Dropbox
DROPBOX_APP_KEY = st.secrets["DROPBOX_APP_KEY"]
DROPBOX_APP_SECRET = st.secrets["DROPBOX_APP_SECRET"]
DROPBOX_REFRESH_TOKEN = st.secrets["DROPBOX_REFRESH_TOKEN"]

dbx = dropbox.Dropbox(
    app_key=DROPBOX_APP_KEY,
    app_secret=DROPBOX_APP_SECRET,
    oauth2_refresh_token=DROPBOX_REFRESH_TOKEN
)

# Configuração das empresas
EMPRESAS = {
    "Magnetic Sky Lda": {
        "path": "/Pedro Couto/Projectos/Alcalá_Arc_Amoreira/Gestão operacional/RH/Processamento Salários Magnetic/Gestão Colaboradores Magnetic.xlsx",
        "seccoes": ["Arc", "Alcalá", "Amoreira TA"]
    },
    "CCM Retail Lda (Pingo Doce)": {
        "path": "/Pedro Couto/Projectos/Pingo Doce/Pingo Doce/2. Operação/1. Recursos Humanos/Processamento salarial/Gestão Colaboradores.xlsx",
        "seccoes": ["Charcutaria/Lacticínios", "Frente de Loja", "Frutas e Vegetais", "Gerência", 
                    "Não Perecíveis (reposição)", "Padaria e Take Away", "Peixaria", "Quiosque", "Talho"]
    }
}

BAIRROS_FISCAIS = [
    "01-AVEIRO - 19-AGUEDA", "01-AVEIRO - 27-ALBERGARIA-A-VELHA", "01-AVEIRO - 35-ANADIA", "01-AVEIRO - 43-AROUCA",
    "01-AVEIRO - 51-AVEIRO-1", "01-AVEIRO - 60-CASTELO DE PAIVA", "01-AVEIRO - 78-ESPINHO", "01-AVEIRO - 86-ESTARREJA",
    "01-AVEIRO - 94-ST. MARIA FEIRA-1", "01-AVEIRO - 108-ILHAVO", "01-AVEIRO - 116-MEALHADA", "01-AVEIRO - 124-MURTOSA",
    "01-AVEIRO - 132-OLIVEIRA AZEMEIS", "01-AVEIRO - 140-OLIVEIRA DO BAIRRO", "01-AVEIRO - 159-OVAR",
    "01-AVEIRO - 167-S. JOAO DA MADEIRA", "01-AVEIRO - 175-SEVER DO VOUGA", "01-AVEIRO - 183-VAGOS",
    "01-AVEIRO - 191-VALE DE CAMBRA", "02-BEJA - 205-ALJUSTREL", "02-BEJA - 213-ALMODOVAR", "02-BEJA - 221-ALVITO",
    "02-BEJA - 230-BARRANCOS", "02-BEJA - 248-BEJA", "02-BEJA - 256-CASTRO VERDE", "02-BEJA - 264-CUBA",
    "02-BEJA - 272-FERREIRA DO ALENTEJO", "02-BEJA - 280-MERTOLA", "02-BEJA - 299-MOURA", "02-BEJA - 302-ODEMIRA",
    "02-BEJA - 310-OURIQUE", "02-BEJA - 329-SERPA", "02-BEJA - 337-VIDIGUEIRA",
    "07-EVORA - 914-EVORA", "12-PORTALEGRE - 1660-ELVAS", "13-PORTO - 1910-VILA NOVA DE GAIA-1",
    "15-SETUBAL - 2232-SETUBAL-1", "11-LISBOA - 3069-LISBOA-1 BAIRRO", "11-LISBOA - 3085-LISBOA-3 BAIRRO",
    "13-PORTO - 3174-PORTO-1 BAIRRO", "13-PORTO - 3204-VILA NOVA DE GAIA-2", "21-PONTA DELGADA - 2992-PONTA DELGADA",
    "22-FUNCHAL - 2810-FUNCHAL-1", "22-FUNCHAL - 2895-SANTANA"
]

def validar_email(email):
    if "@" not in email:
        return False
    partes = email.split("@")
    return len(partes) == 2 and len(partes[0]) > 0 and len(partes[1]) > 0

def validar_nif(nif):
    return len(str(nif).replace(" ", "")) == 9 and str(nif).isdigit()

def validar_niss(niss):
    return len(str(niss).replace(" ", "")) == 11 and str(niss).isdigit()

def validar_telemovel(tel):
    tel_clean = str(tel).replace(" ", "")
    return len(tel_clean) == 9 and tel_clean.isdigit()

def validar_iban(iban):
    iban_clean = iban.replace(" ", "")
    if not iban_clean.startswith("PT50"):
        return False
    if len(iban_clean) != 25:
        return False
    return iban_clean[4:].isdigit()

def validar_cc(cc):
    return len(cc.strip()) > 0

def carregar_dados_dropbox(file_path):
    try:
        _, response = dbx.files_download(file_path)
        data = response.content
        df = pd.read_excel(BytesIO(data), sheet_name="Colaboradores")
        return df
    except Exception:
        colunas = [
            "Nome Completo", "Secção", "Nº Horas/Semana", "E-mail", "Data de Nascimento",
            "NISS", "NIF", "Documento de Identificação", "Validade Documento", "Bairro Fiscal",
            "Estado Civil", "Nº Titulares", "Nº Dependentes", "Morada", "IBAN",
            "Data de Admissão", "Nacionalidade", "Telemóvel", "Subsídio Alimentação Diário",
            "Pessoa com Deficiência", "Tipo IRS", "% IRS Fixa", "Data de Registo"
        ]
        return pd.DataFrame(columns=colunas)

def guardar_dados_dropbox(df, file_path):
    try:
        _, response = dbx.files_download(file_path)
        existing_data = response.content
        wb = load_workbook(BytesIO(existing_data))
        if "Colaboradores" in wb.sheetnames:
            del wb["Colaboradores"]
        ws = wb.create_sheet("Colaboradores")
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        dbx.files_upload(output.read(), file_path, mode=dropbox.files.WriteMode.overwrite)
        return True
    except Exception as e:
        st.error(f"Erro ao guardar no Dropbox: {e}")
        return False

# Interface
st.title("📋 Registo de Colaboradores")
st.markdown("---")

# Seleção da empresa
empresa_selecionada = st.selectbox(
    "🏢 Selecione a Empresa *",
    options=list(EMPRESAS.keys()),
    help="Escolha a empresa para registar o colaborador"
)

st.markdown(f"### Registando para: **{empresa_selecionada}**")
st.markdown("---")

config_empresa = EMPRESAS[empresa_selecionada]

with st.form("formulario_colaborador"):
    st.subheader("Dados Pessoais")
    col1, col2 = st.columns(2)
    
    with col1:
        nome = st.text_input("Nome Completo *", help="Nome completo do colaborador")
        email = st.text_input("E-mail *", help="Email corporativo ou pessoal (deve conter @)")
        data_nascimento = st.date_input(
            "Data de Nascimento *",
            min_value=datetime(1950, 1, 1).date(),
            max_value=datetime.now().date(),
            help="Formato: dd/mm/aaaa"
        )
        nif = st.text_input("NIF *", max_chars=9, help="9 dígitos")
        niss = st.text_input("NISS *", max_chars=11, help="11 dígitos")
    
    with col2:
        telemovel = st.text_input("Telemóvel *", max_chars=9, help="9 dígitos")
        nacionalidade = st.text_input("Nacionalidade *", help="Ex: Portuguesa")
        bairro_fiscal = st.selectbox(
            "Bairro Fiscal *",
            options=BAIRROS_FISCAIS,
            help="Serviço de finanças da área de residência"
        )
        pessoa_deficiencia = st.selectbox(
            "Pessoa com deficiência? *",
            options=["Não", "Sim"],
            help="Para aplicação correta da tabela de IRS"
        )
        doc_identificacao = st.text_input(
            "Documento de Identificação *",
            help="Formato CC: 12345678 0 ZW0 ou 'Passaporte' ou 'Cartão de Residência'"
        )
        validade_doc = st.date_input("Validade do Documento *", help="Formato: dd/mm/aaaa")
    
    st.subheader("Situação Familiar")
    col3, col4 = st.columns(2)
    
    with col3:
        estado_civil = st.selectbox(
            "Estado Civil / Nº Titulares *",
            ["Casado 1", "Casado 2", "Não Casado"],
            help="Casado 1: único titular casado | Casado 2: ambos titulares | Não Casado"
        )
        num_titulares = st.number_input(
            "Nº Titulares *", min_value=1, max_value=2, value=1,
            help="Número de titulares do agregado familiar"
        )
    
    with col4:
        num_dependentes = st.number_input(
            "Nº Dependentes *", min_value=0, value=0,
            help="Número de dependentes a cargo"
        )
    
    st.subheader("Morada")
    morada = st.text_area(
        "Morada Completa *",
        help="Completa com rua, lote, porta, andar, código postal e cidade"
    )
    
    st.subheader("Dados Profissionais")
    col5, col6 = st.columns(2)
    
    with col5:
        secao = st.selectbox(
            "Secção *",
            options=config_empresa["seccoes"],
            help="Departamento ou secção do colaborador"
        )
        horas_semana = st.selectbox(
            "Nº Horas/Semana *",
            [16, 20, 40],
            help="Horas de trabalho semanais (16h, 20h ou 40h)"
        )
        data_admissao = st.date_input("Data de Admissão *", help="Formato: dd/mm/aaaa")
    
    with col6:
        iban = st.text_input(
            "IBAN *",
            max_chars=25,
            placeholder="PT50 0000 0000 0000 0000 0000 0",
            help="Formato: PT50 seguido de 21 dígitos (25 caracteres no total)"
        )
        subsidio_alimentacao = st.number_input(
            "Subsídio de Alimentação Diário *",
            min_value=0.0,
            step=0.01,
            format="%.2f",
            help="Valor diário do subsídio de alimentação em euros"
        )
        tipo_irs = st.selectbox(
            "Tipo de IRS *",
            options=["Automático (por tabela)", "Percentagem fixa"],
            help="Escolha como calcular o IRS deste colaborador"
        )
    
    if tipo_irs == "Percentagem fixa":
        percentagem_irs_fixa = st.number_input(
            "Percentagem IRS Fixa *",
            min_value=0.0,
            max_value=100.0,
            step=0.1,
            format="%.1f",
            help="Percentagem de IRS a aplicar (exemplo: 15.5)"
        )
    else:
        percentagem_irs_fixa = None
    
    st.markdown("---")
    st.caption("* Campos obrigatórios")
    
    submitted = st.form_submit_button("✅ Submeter Registo", use_container_width=True)
    
    if submitted:
        erros = []
        
        if not nome or len(nome) < 3:
            erros.append("Nome completo é obrigatório")
        if not email or not validar_email(email):
            erros.append("Email inválido (deve conter @)")
        if not nif or not validar_nif(nif):
            erros.append("NIF deve ter 9 dígitos")
        if not niss or not validar_niss(niss):
            erros.append("NISS deve ter 11 dígitos")
        if not telemovel or not validar_telemovel(telemovel):
            erros.append("Telemóvel deve ter 9 dígitos")
        if not doc_identificacao or not validar_cc(doc_identificacao):
            erros.append("Documento de identificação em formato inválido")
        if not iban or not validar_iban(iban):
            erros.append("IBAN deve estar no formato PT50 seguido de 21 dígitos")
        if not morada or len(morada) < 10:
            erros.append("Morada completa é obrigatória")
        if not nacionalidade:
            erros.append("Nacionalidade é obrigatória")
        if subsidio_alimentacao <= 0:
            erros.append("Subsídio de alimentação deve ser maior que zero")
        if tipo_irs == "Percentagem fixa" and (percentagem_irs_fixa is None or percentagem_irs_fixa < 0):
            erros.append("Percentagem de IRS fixa deve ser definida e válida")
        
        if erros:
            st.error("Por favor corrija os seguintes erros:")
            for erro in erros:
                st.error(f"• {erro}")
        else:
            novo_registo = {
                "Nome Completo": nome,
                "Secção": secao,
                "Nº Horas/Semana": horas_semana,
                "E-mail": email,
                "Data de Nascimento": data_nascimento.strftime("%d/%m/%Y"),
                "NISS": niss,
                "NIF": nif,
                "Documento de Identificação": doc_identificacao,
                "Validade Documento": validade_doc.strftime("%d/%m/%Y"),
                "Bairro Fiscal": bairro_fiscal,
                "Estado Civil": estado_civil,
                "Nº Titulares": num_titulares,
                "Nº Dependentes": num_dependentes,
                "Morada": morada,
                "IBAN": iban,
                "Data de Admissão": data_admissao.strftime("%d/%m/%Y"),
                "Nacionalidade": nacionalidade,
                "Telemóvel": telemovel,
                "Subsídio Alimentação Diário": subsidio_alimentacao,
                "Pessoa com Deficiência": pessoa_deficiencia,
                "Tipo IRS": tipo_irs,
                "% IRS Fixa": percentagem_irs_fixa if percentagem_irs_fixa is not None else "",
                "Data de Registo": datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            }
            
            with st.spinner("A guardar..."):
                df = carregar_dados_dropbox(config_empresa["path"])
                df = pd.concat([df, pd.DataFrame([novo_registo])], ignore_index=True)
                
                if guardar_dados_dropbox(df, config_empresa["path"]):
                    st.success(f"✅ Registo guardado com sucesso em {empresa_selecionada}!")
                    st.balloons()
                    st.info(f"Total de colaboradores registados em {empresa_selecionada}: {len(df)}")
                else:
                    st.error("❌ Erro ao guardar o registo. Tente novamente.")

st.markdown("---")
st.caption(f"Formulário de Registo de Colaboradores | {empresa_selecionada} | Dados guardados de forma segura no Dropbox")
